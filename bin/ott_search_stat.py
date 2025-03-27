#! /usr/bin/env python3
import os

import openpyxl
import requests

# This script is used to send a OTT search request and feed to result into a excel sheet.
# List of search text are added in queries.txt file
#
# You can call it with python3 bin/ott_search_stat.py
#
# results will be populate into  search_results.xlsx with three different sheets,
# results will be populate into  search_results.xlsx with three different sheets,
# one per each section in front end search result page


def get_goods_nomenclature_item_id(data):
    source = data.get("_source", {})

    # Case 1: Directly available
    if "goods_nomenclature_item_id" in source:
        return source["goods_nomenclature_item_id"]

    # Case 2: Nested inside reference
    reference = source.get("reference", {})
    return reference.get("goods_nomenclature_item_id", "N/A")
    return reference.get("goods_nomenclature_item_id", "N/A")


# Fetch JSON API data
# url = "https://dev.trade-tariff.service.gov.uk"
url = "http://localhost:3000"
endpoint = "/api/v2/search?q="

file_path = "search_results.xlsx"
if os.path.exists(file_path):
    workbook = openpyxl.load_workbook(file_path)
else:
    workbook = openpyxl.Workbook()

workbook.remove(workbook.active)  # Remove default sheet
workbook.create_sheet("Commodity Match")
workbook.create_sheet("Results")
workbook.create_sheet("Other Results")

for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    sheet.append(["Query", "Type", "Commodity 1", "Commodity 2", "Commodity 3", "Commodity 4", "Commodity 5"])


with open("queries.txt", "r") as file:
    queries = [line.strip() for line in file.readlines()]

for query in queries:
    full_url = url + endpoint + query
    response = requests.get(full_url)

    # Check if the request was successful
    if response.status_code == 200:
        print(query)
        json_data = response.json()

        type = json_data["data"]["type"]
        attributes = json_data["data"]["attributes"]
        commodity_codes = []

        if type == "exact_search":
            commodity_codes = [attributes["entry"]["id"], "N/A", "N/A", "N/A", "N/A"]
        else:
            goods_commodities = attributes["goods_nomenclature_match"]["commodities"]
            goods_chapters = attributes["goods_nomenclature_match"]["chapters"]
            goods_headings = attributes["goods_nomenclature_match"]["headings"]
            reference_commodities = attributes["reference_match"]["commodities"]
            reference_chapters = attributes["reference_match"]["chapters"]
            reference_headings = attributes["reference_match"]["headings"]

            all_commodities = sorted(
                goods_commodities + reference_commodities,
                key=lambda x: x["_score"],
                reverse=True,
            )

            all_chapter_heading_results = sorted(
                reference_chapters + reference_headings,
                key=lambda x: x["_score"],
                reverse=True,
            )

            all_chapter_heading_oter_results = sorted(
                goods_chapters + goods_headings, key=lambda x: x["_score"], reverse=True
            )

            commodity_codes = [
                (
                    get_goods_nomenclature_item_id(all_commodities[i])
                    if len(all_commodities) > i
                    else "N/A"
                )
                for i in range(5)
            ]

            results_codes = [
                (
                    all_chapter_heading_results[i]["_source"]["reference"][
                        "goods_nomenclature_item_id"
                    ]
                    if len(all_chapter_heading_results) > i
                    else "N/A"
                )
                for i in range(5)
            ]

            other_results_codes = [
                (
                    all_chapter_heading_oter_results[i]["_source"][
                        "goods_nomenclature_item_id"
                    ]
                    if len(all_chapter_heading_oter_results) > i
                    else "N/A"
                )
                for i in range(5)
            ]

        workbook["Commodity Match"].append([query] + [type] + commodity_codes)
        workbook["Results"].append([query] + [type] + results_codes)
        workbook["Other Results"].append([query] + [type] + other_results_codes)

    else:
        print(f"Query: {query} -> Error: {response.status_code}")

workbook.save(file_path)
